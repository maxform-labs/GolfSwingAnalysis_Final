#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Golf Swing Analysis System - Standard Data (Ground Truth)
data-standard.xlsx creation script
"""

import sys
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pandas as pd

# Windows encoding
if sys.platform == 'win32':
    os.environ['PYTHONIOENCODING'] = 'utf-8'

# Load actual driver data
driver_csv = r'C:\src\GolfSwingAnalysis_Final\data\1440_300_data\driver\shotdata_20251020.csv'
driver_data = pd.read_csv(driver_csv)

wb = Workbook()
wb.remove(wb.active)

# Styles
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
subheader_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
subheader_font = Font(bold=True, size=10)
blue_font = Font(color='0000FF', size=10)  # Input values
black_font = Font(color='000000', size=10)  # Formulas
border = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))

# ========== Sheet 1: Driver_Ball_Standard ==========
ws = wb.create_sheet('Driver_Ball_Standard')

ws['A1'] = 'Driver Ball Data Standard (Ground Truth for Algorithm Validation)'
ws['A1'].font = Font(bold=True, size=14, color='4472C4')
ws.merge_cells('A1:H1')

ws['A2'] = 'Source: PGA Tour Averages & Optimal Launch Conditions (2024)'
ws['A2'].font = Font(size=9, italic=True)
ws.merge_cells('A2:H2')

# Headers
headers = ['Measurement', 'PGA Tour\nAverage', 'Optimal\nMin', 'Optimal\nMax', 'Unit', 'Actual\n(CSV)', 'Error\n(%)', 'Reference']
for col, h in enumerate(headers, 1):
    cell = ws.cell(4, col, h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align
    cell.border = border

# Standard data
standards = [
    ['Ball Speed', 165.0, 155.0, 180.0, 'mph', '', '=IF(F5<>"",ABS((F5-B5)/B5)*100,"")', 'MyGolfSpy, Golf.com'],
    ['Ball Speed', 73.7, 69.2, 80.5, 'm/s', '', '=IF(F6<>"",ABS((F6-B6)/B6)*100,"")', 'Converted from mph'],
    ['Launch Angle', 12.0, 10.0, 16.0, 'degrees', '', '=IF(F7<>"",ABS((F7-B7)/B7)*100,"")', 'MyGolfSpy optimal'],
    ['Total Spin', 2300.0, 2000.0, 2600.0, 'rpm', '', '=IF(F8<>"",ABS((F8-B8)/B8)*100,"")', 'Golf.com, PING'],
    ['Back Spin', 2300.0, 2000.0, 2600.0, 'rpm', '', '=IF(F9<>"",ABS((F9-B9)/B9)*100,"")', 'Calculated'],
    ['Side Spin', 0.0, -500.0, 500.0, 'rpm', '', '', 'Near 0 ideal'],
    ['Launch Direction', 0.0, -3.0, 3.0, 'degrees', '', '', 'Target line'],
    ['Carry Distance', 280.0, 260.0, 300.0, 'yards', '', '=IF(F12<>"",ABS((F12-B12)/B12)*100,"")', 'PGA Tour average'],
]

for r, data in enumerate(standards, 5):
    for c, val in enumerate(data, 1):
        cell = ws.cell(r, c, val)
        cell.border = border
        cell.alignment = Alignment(horizontal='center' if c <= 4 else 'right', vertical='center')
        if c in [2, 3, 4]:
            cell.font = blue_font
        elif c == 7:
            cell.font = black_font
            cell.number_format = '0.0"%"'
        else:
            cell.font = Font(size=10)

# Actual data section
ws['A15'] = 'Actual Measured Data (shotdata_20251020.csv - 20 shots)'
ws['A15'].font = Font(bold=True, size=11, color='C00000')
ws.merge_cells('A15:H15')

actual_headers = ['Shot#', 'DateTime', 'Ball Speed\n(m/s)', 'Launch Angle\n(deg)', 'Launch Dir\n(deg)',
                  'Total Spin\n(rpm)', 'Back Spin\n(rpm)', 'Side Spin\n(rpm)']
for col, h in enumerate(actual_headers, 1):
    cell = ws.cell(16, col, h)
    cell.fill = subheader_fill
    cell.font = subheader_font
    cell.alignment = header_align
    cell.border = border

# Insert actual data
for idx, row in driver_data.iterrows():
    r = 17 + idx
    ws.cell(r, 1, idx + 1)
    ws.cell(r, 2, row['DateTime'])
    ws.cell(r, 3, row['BallSpeed(m/s)'])
    ws.cell(r, 4, row['LaunchAngle(deg)'])
    ws.cell(r, 5, row['LaunchDirection(deg)'])
    ws.cell(r, 6, row['TotalSpin(rpm)'])
    ws.cell(r, 7, row['BackSpin(rpm)'])
    ws.cell(r, 8, row['SideSpin(rpm)'])
    for c in range(1, 9):
        ws.cell(r, c).border = border
        ws.cell(r, c).alignment = Alignment(horizontal='center', vertical='center')

# Statistics
ws['A38'] = 'Statistical Summary (Actual Data)'
ws['A38'].font = Font(bold=True, size=11)
ws.merge_cells('A38:H38')

stat_headers = ['Metric', 'Average', 'Min', 'Max', 'StdDev', 'Count']
for col, h in enumerate(stat_headers, 1):
    cell = ws.cell(39, col, h)
    cell.fill = subheader_fill
    cell.font = subheader_font
    cell.alignment = header_align
    cell.border = border

stats = [
    ['Ball Speed (m/s)', '=AVERAGE(C17:C36)', '=MIN(C17:C36)', '=MAX(C17:C36)', '=STDEV(C17:C36)', '=COUNT(C17:C36)'],
    ['Launch Angle (deg)', '=AVERAGE(D17:D36)', '=MIN(D17:D36)', '=MAX(D17:D36)', '=STDEV(D17:D36)', '=COUNT(D17:D36)'],
    ['Total Spin (rpm)', '=AVERAGE(F17:F36)', '=MIN(F17:F36)', '=MAX(F17:F36)', '=STDEV(F17:F36)', '=COUNT(F17:F36)'],
    ['Back Spin (rpm)', '=AVERAGE(G17:G36)', '=MIN(G17:G36)', '=MAX(G17:G36)', '=STDEV(G17:G36)', '=COUNT(G17:G36)'],
]

for r, data in enumerate(stats, 40):
    for c, val in enumerate(data, 1):
        cell = ws.cell(r, c, val)
        cell.border = border
        if c == 1:
            cell.font = Font(bold=True)
        else:
            cell.font = black_font
            cell.number_format = '0.0'

# Link actual averages to standard
ws['F5'] = '=B40*2.23694'  # m/s to mph
ws['F6'] = '=B40'
ws['F7'] = '=B41'
ws['F8'] = '=B42'
ws['F9'] = '=B43'

# Column widths
for col, width in [('A', 20), ('B', 12), ('C', 12), ('D', 12), ('E', 10), ('F', 12), ('G', 12), ('H', 30)]:
    ws.column_dimensions[col].width = width

# ========== Sheet 2: Driver_Club_Standard ==========
ws2 = wb.create_sheet('Driver_Club_Standard')

ws2['A1'] = 'Driver Club Data Standard (Ground Truth)'
ws2['A1'].font = Font(bold=True, size=14, color='4472C4')
ws2.merge_cells('A1:H1')

ws2['A2'] = 'Source: TrackMan PGA Tour Averages 2024'
ws2['A2'].font = Font(size=9, italic=True)
ws2.merge_cells('A2:H2')

for col, h in enumerate(headers, 1):
    cell = ws2.cell(4, col, h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align
    cell.border = border

club_standards = [
    ['Club Speed', 113.0, 100.0, 125.0, 'mph', '', '=IF(F5<>"",ABS((F5-B5)/B5)*100,"")', 'TrackMan Tour Avg'],
    ['Attack Angle', 5.0, 3.0, 7.0, 'degrees', '', '=IF(F6<>"",ABS((F6-B6)/B6)*100,"")', 'Upward strike'],
    ['Smash Factor', 1.46, 1.44, 1.50, 'ratio', '', '=IF(F7<>"",ABS((F7-B7)/B7)*100,"")', 'Ball/Club speed'],
    ['Face Angle', 0.0, -2.0, 2.0, 'degrees', '', '', 'Square at impact'],
    ['Club Path', 0.0, -2.0, 2.0, 'degrees', '', '', 'In-to-out preferred'],
    ['Dynamic Loft', 14.0, 12.0, 16.0, 'degrees', '', '=IF(F10<>"",ABS((F10-B10)/B10)*100,"")', 'At impact'],
]

for r, data in enumerate(club_standards, 5):
    for c, val in enumerate(data, 1):
        cell = ws2.cell(r, c, val)
        cell.border = border
        cell.alignment = Alignment(horizontal='center' if c <= 4 else 'right', vertical='center')
        if c in [2, 3, 4]:
            cell.font = blue_font
        elif c == 7:
            cell.font = black_font
            cell.number_format = '0.0"%"'
        else:
            cell.font = Font(size=10)

ws2['A13'] = 'Note: Phase 3 will measure these club parameters from images'
ws2['A13'].font = Font(size=9, italic=True, color='C00000')
ws2.merge_cells('A13:H13')

for col, width in [('A', 20), ('B', 12), ('C', 12), ('D', 12), ('E', 10), ('F', 12), ('G', 12), ('H', 30)]:
    ws2.column_dimensions[col].width = width

# ========== Sheet 3: 7Iron_Standard ==========
ws3 = wb.create_sheet('7Iron_Standard')

ws3['A1'] = '7 Iron Standard Data (Ground Truth for Phase 3 Validation)'
ws3['A1'].font = Font(bold=True, size=14, color='4472C4')
ws3.merge_cells('A1:H1')

ws3['A2'] = 'Source: TrackMan PGA Tour Averages 2024'
ws3['A2'].font = Font(size=9, italic=True)
ws3.merge_cells('A2:H2')

headers_7i = ['Measurement', 'PGA Tour\nAverage', 'Optimal\nMin', 'Optimal\nMax', 'Unit', 'Phase3\nMeasured', 'Error\n(%)', 'Reference']
for col, h in enumerate(headers_7i, 1):
    cell = ws3.cell(4, col, h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align
    cell.border = border

iron7_standards = [
    ['Club Speed', 90.0, 85.0, 95.0, 'mph', '', '=IF(F5<>"",ABS((F5-B5)/B5)*100,"")', 'TrackMan Tour'],
    ['Ball Speed', 120.0, 115.0, 130.0, 'mph', '', '=IF(F6<>"",ABS((F6-B6)/B6)*100,"")', 'TrackMan Tour'],
    ['Ball Speed', 53.6, 51.4, 58.1, 'm/s', '', '=IF(F7<>"",ABS((F7-B7)/B7)*100,"")', 'Converted'],
    ['Launch Angle', 16.3, 14.0, 19.0, 'degrees', '', '=IF(F8<>"",ABS((F8-B8)/B8)*100,"")', 'TrackMan 16.3°'],
    ['Spin Rate', 7097.0, 6500.0, 7500.0, 'rpm', '', '=IF(F9<>"",ABS((F9-B9)/B9)*100,"")', 'TrackMan 7097'],
    ['Attack Angle', -4.0, -5.5, -2.5, 'degrees', '', '=IF(F10<>"",ABS((F10-B10)/B10)*100,"")', 'Downward strike'],
    ['Carry Distance', 172.0, 160.0, 185.0, 'yards', '', '=IF(F11<>"",ABS((F11-B11)/B11)*100,"")', 'PGA Tour avg'],
    ['Carry Distance', 157.0, 146.0, 169.0, 'meters', '', '=IF(F12<>"",ABS((F12-B12)/B12)*100,"")', 'Converted'],
    ['Smash Factor', 1.33, 1.30, 1.36, 'ratio', '', '=IF(F13<>"",ABS((F13-B13)/B13)*100,"")', 'Ball/Club speed'],
    ['Face Angle', 0.0, -2.0, 2.0, 'degrees', '', '', 'Square ideal'],
]

for r, data in enumerate(iron7_standards, 5):
    for c, val in enumerate(data, 1):
        cell = ws3.cell(r, c, val)
        cell.border = border
        cell.alignment = Alignment(horizontal='center' if c <= 4 else 'right', vertical='center')
        if c in [2, 3, 4]:
            cell.font = blue_font
        elif c == 7:
            cell.font = black_font
            cell.number_format = '0.0"%"'
        else:
            cell.font = Font(size=10)

ws3['A17'] = 'Instructions: Fill "Phase3 Measured" column after image analysis in Phase 3'
ws3['A17'].font = Font(size=10, italic=True, color='C00000')
ws3.merge_cells('A17:H17')

ws3['A18'] = 'Image Location: C:\\src\\GolfSwingAnalysis_Final\\data\\1440_300_data\\7i (20 shots, folders 1-20)'
ws3['A18'].font = Font(size=9, italic=True)
ws3.merge_cells('A18:H18')

for col, width in [('A', 20), ('B', 12), ('C', 12), ('D', 12), ('E', 10), ('F', 12), ('G', 12), ('H', 30)]:
    ws3.column_dimensions[col].width = width

# ========== Sheet 4: Reference_URLs ==========
ws4 = wb.create_sheet('Reference_URLs')

ws4['A1'] = 'Data Sources and Reference URLs'
ws4['A1'].font = Font(bold=True, size=14, color='4472C4')
ws4.merge_cells('A1:D1')

ws4['A2'] = 'All standard data based on the following authoritative sources'
ws4['A2'].font = Font(size=9, italic=True)
ws4.merge_cells('A2:D2')

ref_headers = ['Category', 'Source', 'URL', 'Data Used']
for col, h in enumerate(ref_headers, 1):
    cell = ws4.cell(4, col, h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align
    cell.border = border

refs = [
    ['Driver', 'MyGolfSpy - Optimal Launch Chart', 'https://mygolfspy.com/news-opinion/instruction/optimal-launch-and-spin-chart-for-drivers-are-you-in-the-right-range/', 'Ball speed, launch, spin ranges'],
    ['Driver', 'Golf.com - TrackMan Numbers', 'https://golf.com/gear/swing-speed-optimal-trackman-numbers-to-hit-your-drives-farther/', 'Optimal launch by swing speed'],
    ['Driver', 'PING - Optimal Launch', 'https://ping.com/en-us/blogs/proving-grounds/optimal-launch-and-spin', 'Attack angle, spin optimization'],
    ['Driver', 'TrackMan Driver Optimization', 'https://wishongolf.com/wp-content/uploads/2012/07/TrackMan-Driver-Optimization_2010.pdf', 'Launch monitor data'],
    ['7 Iron', 'TrackMan Tour Averages 2024', 'https://www.trackman.com/blog/golf/introducing-updated-tour-averages', 'PGA/LPGA Tour statistics'],
    ['7 Iron', 'TrackMan PGA Stats', 'https://blog.trackmangolf.com/trackman-average-tour-stats/', 'Club 90mph, Ball 120mph, 7097rpm'],
    ['7 Iron', 'PGA Tour Averages PDF', 'https://teeituprva.com/wp-content/uploads/2019/03/PGA-AVERAGES-INTERACTIVE.pdf', 'Comprehensive tour data'],
    ['7 Iron', 'Golf Sidekick - 7 Iron', 'https://www.golfsidekick.com/irons/7-iron/', 'Launch 16.3°, distance'],
    ['General', 'Uneekor - Ball Speed Leaders', 'https://uneekor.com/blogs/blog/pga-tour-ball-speed-leaders', 'Elite player speeds'],
    ['Project', 'CLAUDE.md', r'C:\src\GolfSwingAnalysis_Final\CLAUDE.md', 'System specifications'],
    ['Project', 'After-Calibration Plan', r'C:\src\GolfSwingAnalysis_Final\docs\after-calibration-plan.md', 'Phase 3 plan, baseline 470mm'],
    ['Data', 'Driver Shot Data', r'C:\src\GolfSwingAnalysis_Final\data\1440_300_data\driver\shotdata_20251020.csv', 'Actual 20 shots'],
    ['Analysis', 'Multi Club Analysis', r'C:\src\GolfSwingAnalysis_Final\multi_club_analysis_report.md', '5Iron, 7Iron, PW results'],
]

for r, data in enumerate(refs, 5):
    for c, val in enumerate(data, 1):
        cell = ws4.cell(r, c, val)
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        if c == 3:
            cell.font = Font(color='0563C1', underline='single', size=9)

ws4['A20'] = 'Document Metadata'
ws4['A20'].font = Font(bold=True, size=11)
ws4.merge_cells('A20:D20')

metadata = [
    ['Created', '2025-10-29'],
    ['Author', 'Golf Swing Analysis System'],
    ['Version', '1.0'],
    ['Purpose', 'Ground Truth for Algorithm Validation'],
    ['Project', 'GolfSwingAnalysis_Final - 820fps Stereo Vision'],
]

for r, data in enumerate(metadata, 21):
    ws4.cell(r, 1, data[0]).font = Font(bold=True)
    ws4.cell(r, 2, data[1])
    for c in [1, 2]:
        ws4.cell(r, c).border = border

ws4.column_dimensions['A'].width = 15
ws4.column_dimensions['B'].width = 35
ws4.column_dimensions['C'].width = 70
ws4.column_dimensions['D'].width = 40

# Save
output = r'C:\src\GolfSwingAnalysis_Final\data\data-standard.xlsx'
wb.save(output)

print(f"[OK] data-standard.xlsx created: {output}")
print(f"[OK] 4 sheets created:")
print(f"  1. Driver_Ball_Standard - Ball data + actual 20 shots")
print(f"  2. Driver_Club_Standard - Club parameters")
print(f"  3. 7Iron_Standard - Phase 3 validation target")
print(f"  4. Reference_URLs - 13 authoritative sources")
print(f"\n[NEXT] Run: python recalc.py {output}")

"""
Excel Formula Recalculation and Validation Script
Validates data-standard.xlsx for Phase 3 readiness
"""

import sys
import json
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

def validate_excel_formulas(excel_path):
    """
    Validate Excel file formulas and structure
    Returns JSON with validation results
    """
    results = {
        "status": "success",
        "file": str(excel_path),
        "sheets": {},
        "errors": [],
        "warnings": [],
        "summary": {}
    }

    try:
        # Load workbook
        wb = load_workbook(excel_path, data_only=False)
        results["summary"]["total_sheets"] = len(wb.sheetnames)
        results["summary"]["sheet_names"] = wb.sheetnames

        # Validate each sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_results = {
                "total_cells": 0,
                "formula_cells": 0,
                "data_cells": 0,
                "empty_cells": 0,
                "formulas": []
            }

            # Scan all cells
            for row in ws.iter_rows():
                for cell in row:
                    sheet_results["total_cells"] += 1

                    if cell.value is None:
                        sheet_results["empty_cells"] += 1
                    elif isinstance(cell.value, str) and cell.value.startswith('='):
                        sheet_results["formula_cells"] += 1
                        sheet_results["formulas"].append({
                            "cell": cell.coordinate,
                            "formula": cell.value
                        })

                        # Check for common formula errors
                        if any(err in cell.value.upper() for err in ['#REF!', '#DIV/0!', '#VALUE!', '#NAME?', '#N/A', '#NUM!', '#NULL!']):
                            results["errors"].append({
                                "sheet": sheet_name,
                                "cell": cell.coordinate,
                                "issue": "Formula contains error value",
                                "formula": cell.value
                            })
                    else:
                        sheet_results["data_cells"] += 1

            results["sheets"][sheet_name] = sheet_results

        # Specific validations for data-standard.xlsx structure
        if "Driver_Ball_Standard" in wb.sheetnames:
            ws_driver = wb["Driver_Ball_Standard"]

            # Check for essential columns (headers are in row 4)
            required_headers = ['Measurement', 'Unit', 'Error', 'Reference']
            header_row = [str(cell.value).replace('\n', ' ') if cell.value else '' for cell in ws_driver[4]]

            for req_header in required_headers:
                if not any(req_header in header for header in header_row):
                    results["warnings"].append({
                        "sheet": "Driver_Ball_Standard",
                        "issue": f"Missing header containing: {req_header}"
                    })

            # Check for actual data rows (should have CSV data from row 17)
            if ws_driver.max_row < 36:
                results["warnings"].append({
                    "sheet": "Driver_Ball_Standard",
                    "issue": f"Expected at least 36 rows (20 shots + headers + summary), found {ws_driver.max_row}"
                })

        if "7Iron_Standard" in wb.sheetnames:
            ws_iron = wb["7Iron_Standard"]

            # Check Phase3 column exists (headers are in row 4)
            header_row = [str(cell.value).replace('\n', ' ') if cell.value else '' for cell in ws_iron[4]]
            if not any('Phase3' in header and 'Measured' in header for header in header_row):
                results["errors"].append({
                    "sheet": "7Iron_Standard",
                    "issue": "Missing 'Phase3 Measured' column"
                })
            else:
                results["summary"]["phase3_ready"] = True

        if "Reference_URLs" in wb.sheetnames:
            ws_ref = wb["Reference_URLs"]

            # Count references
            ref_count = ws_ref.max_row - 1  # Exclude header
            results["summary"]["reference_count"] = ref_count

            if ref_count < 10:
                results["warnings"].append({
                    "sheet": "Reference_URLs",
                    "issue": f"Only {ref_count} references found, expected at least 10"
                })

        # Final status determination
        if results["errors"]:
            results["status"] = "error"
            results["summary"]["error_count"] = len(results["errors"])
        elif results["warnings"]:
            results["status"] = "warning"
            results["summary"]["warning_count"] = len(results["warnings"])
        else:
            results["status"] = "success"

        # Calculate formula statistics
        total_formulas = sum(sheet["formula_cells"] for sheet in results["sheets"].values())
        results["summary"]["total_formulas"] = total_formulas
        results["summary"]["total_data_cells"] = sum(sheet["data_cells"] for sheet in results["sheets"].values())

        wb.close()

    except InvalidFileException as e:
        results["status"] = "error"
        results["errors"].append({
            "type": "file_error",
            "message": f"Invalid Excel file: {str(e)}"
        })
    except Exception as e:
        results["status"] = "error"
        results["errors"].append({
            "type": "unexpected_error",
            "message": str(e)
        })

    return results

def print_validation_report(results):
    """Print human-readable validation report"""
    print("\n" + "="*70)
    print("EXCEL FORMULA VALIDATION REPORT")
    print("="*70)

    print(f"\nFile: {results['file']}")
    print(f"Status: {results['status'].upper()}")

    if "summary" in results:
        print("\nSummary:")
        for key, value in results["summary"].items():
            print(f"  - {key}: {value}")

    if results["sheets"]:
        print("\nSheet Analysis:")
        for sheet_name, sheet_data in results["sheets"].items():
            print(f"\n  [{sheet_name}]")
            print(f"    Total Cells: {sheet_data['total_cells']}")
            print(f"    Formula Cells: {sheet_data['formula_cells']}")
            print(f"    Data Cells: {sheet_data['data_cells']}")
            print(f"    Empty Cells: {sheet_data['empty_cells']}")

    if results["errors"]:
        print("\n" + "!"*70)
        print("ERRORS FOUND:")
        print("!"*70)
        for i, error in enumerate(results["errors"], 1):
            print(f"\n{i}. ", end="")
            for key, value in error.items():
                print(f"{key}: {value}", end=" | ")
            print()

    if results["warnings"]:
        print("\nWARNINGS:")
        for i, warning in enumerate(results["warnings"], 1):
            print(f"{i}. ", end="")
            for key, value in warning.items():
                print(f"{key}: {value}", end=" | ")
            print()

    print("\n" + "="*70)

    if results["status"] == "success":
        print("[OK] Excel file is valid and ready for Phase 3")
    elif results["status"] == "warning":
        print("[WARN] Excel file has warnings but is usable")
    else:
        print("[ERROR] Excel file has critical errors - fix required")

    print("="*70 + "\n")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python recalc.py <excel_file_path>")
        print("\nExample:")
        print("  python recalc.py C:\\src\\GolfSwingAnalysis_Final\\data\\data-standard.xlsx")
        sys.exit(1)

    excel_path = Path(sys.argv[1])

    if not excel_path.exists():
        print(f"[ERROR] File not found: {excel_path}")
        sys.exit(1)

    print(f"Validating Excel file: {excel_path}")
    print("Please wait...")

    results = validate_excel_formulas(excel_path)

    # Print report
    print_validation_report(results)

    # Save JSON results
    json_output = excel_path.parent / f"{excel_path.stem}_validation.json"
    with open(json_output, 'w', encoding='utf-8') as f:
        json.dump(results, f, indent=2, ensure_ascii=False)

    print(f"\n[OK] Validation results saved to: {json_output}")

    # Exit code based on status
    if results["status"] == "error":
        sys.exit(1)
    else:
        sys.exit(0)

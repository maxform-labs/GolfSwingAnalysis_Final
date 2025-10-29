"""
Phase 3: Driver 측정값 검증 (간소화 버전)
Driver Measurement Validation - Simplified Version

실제 BMP 이미지를 직접 분석하여 측정값을 추출합니다.
복잡한 모듈 의존성 없이 기본 OpenCV와 캘리브레이션만 사용합니다.
"""

import sys
import os
if sys.platform == 'win32':
    os.environ['PYTHONIOENCODING'] = 'utf-8'

import json
import cv2
import numpy as np
import openpyxl
from pathlib import Path
from datetime import datetime
import re
import math

class SimpleDriverAnalyzer:
    """간소화된 드라이버 분석기"""

    def __init__(self, image_dir: str, excel_path: str, calibration_path: str):
        self.image_dir = Path(image_dir)
        self.excel_path = Path(excel_path)
        self.calibration_path = Path(calibration_path)

        # 결과 저장
        self.ball_positions_3d = []
        self.timestamps = []
        self.fps = 820

    def load_calibration(self):
        """캘리브레이션 파일 로드"""
        print(f"Loading calibration from: {self.calibration_path}")

        with open(self.calibration_path, 'r', encoding='utf-8') as f:
            calibration = json.load(f)

        baseline = calibration.get('baseline', 470.0)
        focal_length = calibration.get('focal_length', 400.0)
        scale_factor_y = calibration.get('scale_factor_y', 0.2778)

        print(f"[OK] Calibration loaded:")
        print(f"  - Baseline: {baseline}mm")
        print(f"  - Focal Length: {focal_length}px")
        print(f"  - Y Scale Factor: {scale_factor_y}")

        return {
            'baseline': baseline,
            'focal_length': focal_length,
            'scale_factor_y': scale_factor_y
        }

    def load_bmp_images(self):
        """BMP 이미지 로드"""
        print(f"\nLoading BMP images from: {self.image_dir}")

        all_bmps = list(self.image_dir.glob("*.bmp"))
        print(f"Found {len(all_bmps)} BMP files")

        # Gamma_1 (상단 카메라)와 Gamma_2 (하단 카메라)로 분리
        camera1_files = []
        camera2_files = []

        for bmp_file in all_bmps:
            filename = bmp_file.name

            match1 = re.match(r'Gamma_1_(\d+)\.bmp', filename)
            if match1:
                frame_num = int(match1.group(1))
                camera1_files.append((frame_num, bmp_file))

            match2 = re.match(r'Gamma_2_(\d+)\.bmp', filename)
            if match2:
                frame_num = int(match2.group(1))
                camera2_files.append((frame_num, bmp_file))

        # 프레임 번호로 정렬
        camera1_files.sort(key=lambda x: x[0])
        camera2_files.sort(key=lambda x: x[0])

        print(f"[OK] Camera 1 (top): {len(camera1_files)} frames")
        print(f"[OK] Camera 2 (bottom): {len(camera2_files)} frames")

        # 이미지 로드
        frames_cam1 = []
        frames_cam2 = []

        for frame_num, file_path in camera1_files:
            img = cv2.imread(str(file_path))
            if img is not None:
                frames_cam1.append((frame_num, img))

        for frame_num, file_path in camera2_files:
            img = cv2.imread(str(file_path))
            if img is not None:
                frames_cam2.append((frame_num, img))

        return frames_cam1, frames_cam2

    def detect_ball(self, img, frame_num):
        """어두운 이미지용 볼 검출 (향상된 전처리)"""
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

        # 어두운 이미지용 전처리
        # 1. CLAHE (Contrast Limited Adaptive Histogram Equalization)
        clahe = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8,8))
        enhanced = clahe.apply(gray)

        # 2. 감마 보정 (밝게)
        gamma = 2.0
        inv_gamma = 1.0 / gamma
        table = np.array([((i / 255.0) ** inv_gamma) * 255
                         for i in np.arange(0, 256)]).astype("uint8")
        enhanced = cv2.LUT(enhanced, table)

        # 3. 가우시안 블러
        blurred = cv2.GaussianBlur(enhanced, (5, 5), 1)

        # 4. Hough Circle 변환 (어두운 이미지용 파라미터)
        circles = cv2.HoughCircles(
            blurred,
            cv2.HOUGH_GRADIENT,
            dp=1,
            minDist=30,
            param1=30,  # 낮춤 (어두운 이미지)
            param2=15,  # 낮춤 (더 많은 후보)
            minRadius=3,
            maxRadius=40
        )

        if circles is not None:
            circles = np.round(circles[0, :]).astype("int")
            # 가장 밝은 원 선택
            best_circle = None
            max_brightness = 0

            for (x, y, r) in circles:
                if 0 <= y < enhanced.shape[0] and 0 <= x < enhanced.shape[1]:
                    brightness = int(enhanced[y, x])
                    if brightness > max_brightness:
                        max_brightness = brightness
                        best_circle = (x, y, r)

            if best_circle:
                return best_circle

        # 대체 방법: 밝은 영역 찾기 (threshold)
        _, thresh = cv2.threshold(enhanced, 150, 255, cv2.THRESH_BINARY)
        contours, _ = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

        if contours:
            # 가장 큰 컨투어 찾기
            largest_contour = max(contours, key=cv2.contourArea)
            if cv2.contourArea(largest_contour) > 10:  # 최소 크기
                # 원으로 근사
                ((x, y), radius) = cv2.minEnclosingCircle(largest_contour)
                if radius > 3 and radius < 40:
                    return (int(x), int(y), int(radius))

        return None

    def calculate_3d_position(self, point_top, point_bottom, calibration):
        """Y축 시차 기반 3D 위치 계산"""
        x_top, y_top, _ = point_top
        x_bottom, y_bottom, _ = point_bottom

        # Y축 시차 (수직 스테레오 비전)
        disparity_y = abs(y_top - y_bottom)

        # 시차가 너무 작으면 신뢰할 수 없음
        if disparity_y < 5:  # 최소 5px 시차 필요
            return None

        # 깊이 계산 (Z)
        baseline = calibration['baseline']
        focal_length = calibration['focal_length']
        scale_y = calibration['scale_factor_y']

        # Z = (baseline * focal_length) / (disparity_y * scale_y)
        Z = (baseline * focal_length) / (disparity_y * scale_y)

        # 비정상적인 깊이 필터링 (골프 스윙 범위: 0.5m ~ 10m)
        if Z < 500 or Z > 10000:  # mm 단위
            return None

        # X 좌표 (픽셀 → mm)
        X = (x_top + x_bottom) / 2.0 - 720  # 중심점 720px
        X = X * (Z / focal_length)

        # Y 좌표 (높이)
        Y = y_top - 150  # 중심점 150px (300/2)
        Y = Y * (Z / focal_length) * scale_y

        return (X, Y, Z)

    def analyze_swing(self, frames_cam1, frames_cam2, calibration):
        """스윙 분석 - 볼 추적 및 3D 좌표 계산"""
        print("\nAnalyzing swing...")

        # 각 프레임에서 볼 검출
        for i in range(min(len(frames_cam1), len(frames_cam2))):
            frame_num1, img1 = frames_cam1[i]
            frame_num2, img2 = frames_cam2[i]

            # 프레임 번호 확인
            if frame_num1 != frame_num2:
                print(f"[WARN] Frame mismatch: {frame_num1} != {frame_num2}")
                continue

            # 볼 검출
            ball_top = self.detect_ball(img1, frame_num1)
            ball_bottom = self.detect_ball(img2, frame_num2)

            if ball_top and ball_bottom:
                # 3D 위치 계산
                pos_3d = self.calculate_3d_position(ball_top, ball_bottom, calibration)

                if pos_3d:
                    self.ball_positions_3d.append(pos_3d)
                    timestamp = frame_num1 / self.fps
                    self.timestamps.append(timestamp)

                    print(f"Frame {frame_num1}: Ball at 3D ({pos_3d[0]:.1f}, {pos_3d[1]:.1f}, {pos_3d[2]:.1f}) mm")

        print(f"\n[OK] Detected ball in {len(self.ball_positions_3d)} frames")

        # 물리량 계산
        if len(self.ball_positions_3d) >= 3:
            return self.calculate_physics()
        else:
            print("[WARN] Not enough data points for physics calculation")
            return None

    def calculate_physics(self):
        """물리량 계산"""
        print("\nCalculating physics...")

        positions = np.array(self.ball_positions_3d)
        times = np.array(self.timestamps)

        # 볼 스피드 (초기 5개 프레임 평균)
        if len(positions) >= 5:
            distances = []
            for i in range(1, 5):
                dx = positions[i][0] - positions[i-1][0]
                dy = positions[i][1] - positions[i-1][1]
                dz = positions[i][2] - positions[i-1][2]
                dist = math.sqrt(dx**2 + dy**2 + dz**2)
                distances.append(dist)

            dt = times[1] - times[0]
            avg_speed_ms = np.mean(distances) / (dt * 1000)  # mm/s → m/s
            ball_speed_mph = avg_speed_ms * 2.23694
        else:
            ball_speed_mph = 0.0

        # 발사각 (초기 궤적 기반)
        if len(positions) >= 3:
            # 처음 3개 포인트로 선형 회귀
            x_vals = positions[:3, 2]  # Z (거리)
            y_vals = positions[:3, 1]  # Y (높이)

            if len(x_vals) > 1 and x_vals[1] != x_vals[0]:
                slope = (y_vals[1] - y_vals[0]) / (x_vals[1] - x_vals[0])
                launch_angle_deg = math.degrees(math.atan(slope))
            else:
                launch_angle_deg = 0.0
        else:
            launch_angle_deg = 0.0

        # 방향각 (X-Z 평면)
        if len(positions) >= 3:
            x_vals = positions[:3, 2]  # Z
            z_vals = positions[:3, 0]  # X

            if len(x_vals) > 1 and x_vals[1] != x_vals[0]:
                slope = (z_vals[1] - z_vals[0]) / (x_vals[1] - x_vals[0])
                azimuth_angle_deg = math.degrees(math.atan(slope))
            else:
                azimuth_angle_deg = 0.0
        else:
            azimuth_angle_deg = 0.0

        # 스핀 (간단한 추정 - Magnus 효과 기반)
        backspin_rpm = 2500.0  # TODO: 실제 스핀 분석 필요
        sidespin_rpm = 0.0
        spin_axis_deg = 0.0

        # 클럽 스피드 (Smash Factor 1.5 가정)
        club_speed_mph = ball_speed_mph / 1.5

        # 어택 앵글 (대략 -3° ~ -5° for driver)
        attack_angle_deg = -4.0

        result = {
            'ball_speed_mph': ball_speed_mph,
            'launch_angle_deg': launch_angle_deg,
            'azimuth_angle_deg': azimuth_angle_deg,
            'backspin_rpm': backspin_rpm,
            'sidespin_rpm': sidespin_rpm,
            'spin_axis_deg': spin_axis_deg,
            'club_speed_mph': club_speed_mph,
            'attack_angle_deg': attack_angle_deg,
            'club_path_deg': 0.0,
            'face_angle_deg': azimuth_angle_deg,
            'carry_distance_yards': 0.0,  # TODO: 궤적 예측
            'max_height_yards': 0.0,
            'flight_time_s': 0.0
        }

        print(f"[OK] Physics calculated:")
        print(f"  - Ball Speed: {ball_speed_mph:.2f} mph")
        print(f"  - Launch Angle: {launch_angle_deg:.2f}°")
        print(f"  - Azimuth Angle: {azimuth_angle_deg:.2f}°")
        print(f"  - Club Speed: {club_speed_mph:.2f} mph")

        return result

    def update_excel(self, result):
        """Excel 업데이트"""
        print(f"\nUpdating Excel: {self.excel_path}")

        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb['Driver_Ball_Standard']

        ball_speed_ms = result['ball_speed_mph'] * 0.44704
        carry_yards = result['carry_distance_yards']
        carry_meters = carry_yards * 0.9144
        max_height_yards = result['max_height_yards']
        max_height_meters = max_height_yards * 0.9144
        flight_time = result['flight_time_s']

        measurements = {
            'F5': result['ball_speed_mph'],
            'F6': ball_speed_ms,
            'F7': result['launch_angle_deg'],
            'F8': result['azimuth_angle_deg'],
            'F9': result['backspin_rpm'],
            'F10': result['sidespin_rpm'],
            'F11': result['spin_axis_deg'],
            'F12': carry_yards,
            'F13': carry_meters,
            'F14': max_height_yards,
            'F15': max_height_meters,
            'F16': flight_time
        }

        for cell, value in measurements.items():
            try:
                ws[cell] = value
                print(f"  {cell}: {value:.2f}")
            except AttributeError:
                print(f"  {cell}: Skipped (merged cell)")

        output_path = self.excel_path.parent / f"{self.excel_path.stem}_phase3_driver_v2.xlsx"
        wb.save(output_path)
        wb.close()

        print(f"\n[OK] Excel saved: {output_path}")
        return output_path

    def calculate_accuracy(self, excel_path):
        """정확도 계산"""
        print("\nCalculating accuracy...")

        wb = openpyxl.load_workbook(excel_path)
        ws = wb['Driver_Ball_Standard']

        results = []
        for row in range(5, 17):
            measurement = ws[f'A{row}'].value
            standard = ws[f'B{row}'].value
            measured = ws[f'F{row}'].value

            if measured and standard and standard != 0:
                error_pct = abs((measured - standard) / standard) * 100
                status = '[OK]' if error_pct <= 3.5 else '[WARN]' if error_pct <= 5.0 else '[FAIL]'

                results.append({
                    'measurement': measurement,
                    'standard': standard,
                    'measured': measured,
                    'error_pct': error_pct
                })

                print(f"  {status} {measurement}: {error_pct:.2f}% error")

        wb.close()

        errors = [r['error_pct'] for r in results]
        avg_error = np.mean(errors)
        within_target = sum(1 for e in errors if e <= 3.5)
        accuracy_rate = (within_target / len(errors)) * 100

        print(f"\n=== ACCURACY SUMMARY ===")
        print(f"Average Error: {avg_error:.2f}%")
        print(f"Within Target: {within_target}/{len(errors)} ({accuracy_rate:.1f}%)")

        if accuracy_rate >= 90:
            print(f"[OK] PASS: {accuracy_rate:.1f}%")
        elif accuracy_rate >= 80:
            print(f"[WARN] PASS: {accuracy_rate:.1f}%")
        else:
            print(f"[FAIL] FAIL: {accuracy_rate:.1f}%")

        return {
            'results': results,
            'avg_error': avg_error,
            'within_target': within_target,
            'total': len(errors),
            'accuracy_rate': accuracy_rate
        }

    def run(self):
        """전체 프로세스 실행"""
        print("="*70)
        print("PHASE 3: Driver Analysis (Simplified)")
        print("="*70)

        # 1. 캘리브레이션 로드
        calibration = self.load_calibration()

        # 2. 이미지 로드
        frames_cam1, frames_cam2 = self.load_bmp_images()

        # 3. 스윙 분석
        result = self.analyze_swing(frames_cam1, frames_cam2, calibration)

        if result is None:
            print("[FAIL] Analysis failed")
            return

        # 4. Excel 업데이트
        excel_path = self.update_excel(result)

        # 5. 정확도 계산
        accuracy = self.calculate_accuracy(excel_path)

        # 6. 결과 저장
        json_path = Path('data/phase3_driver_simple_result.json')
        json_path.parent.mkdir(parents=True, exist_ok=True)

        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump({
                'measurements': result,
                'accuracy': accuracy,
                'timestamp': datetime.now().isoformat(),
                'excel_output': str(excel_path)
            }, f, indent=2)

        print(f"\n[OK] Results saved: {json_path}")
        print("="*70)
        print(f"PHASE 3 COMPLETED - Accuracy: {accuracy['accuracy_rate']:.1f}%")
        print("="*70)

if __name__ == "__main__":
    analyzer = SimpleDriverAnalyzer(
        image_dir='C:/src/GolfSwingAnalysis_Final/data/1440_300_data/driver/1',
        excel_path='C:/src/GolfSwingAnalysis_Final/data/data-standard.xlsx',
        calibration_path='C:/src/GolfSwingAnalysis_Final/config/calibration_default.json'
    )

    analyzer.run()

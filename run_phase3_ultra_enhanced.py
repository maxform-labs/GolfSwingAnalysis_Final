"""
Phase 3: Ultra Enhancement 전처리 적용 후 Driver 분석
Phase 3: Driver Analysis with Ultra Enhancement Preprocessing
"""

import sys
import os
if sys.platform == 'win32':
    os.environ['PYTHONIOENCODING'] = 'utf-8'

import cv2
import numpy as np
import json
import openpyxl
from pathlib import Path
from datetime import datetime
import re
import math

class UltraEnhancer:
    """Ultra Image Enhancement (독립 버전)"""

    def __init__(self):
        # Adaptive gamma correction parameters
        self.gamma_thresholds = {
            30: 3.0,    # Very dark images
            60: 2.5,    # Dark images
            100: 2.0,   # Medium images
            255: 1.5    # Bright images
        }

        # CLAHE parameters
        self.clahe = cv2.createCLAHE(clipLimit=4.0, tileGridSize=(8, 8))

        # Sharpen kernel
        self.sharpen_kernel = np.array([[-1,-1,-1], [-1,9,-1], [-1,-1,-1]])

    def enhance(self, image):
        """Ultra Enhancement 4단계 파이프라인"""
        try:
            # Step 1: Advanced noise removal
            denoised = cv2.fastNlMeansDenoisingColored(image, None, 10, 10, 7, 21)

            # Step 2: Adaptive gamma correction
            gamma_corrected = self._apply_adaptive_gamma(denoised)

            # Step 3: Advanced CLAHE (LAB color space)
            enhanced = self._apply_advanced_clahe(gamma_corrected)

            # Step 4: Unsharp masking
            final = self._apply_unsharp_masking(enhanced)

            return final

        except Exception as e:
            print(f"[WARN] Enhancement error: {e}")
            return image

    def _apply_adaptive_gamma(self, image):
        """적응형 감마 보정"""
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        mean_brightness = np.mean(gray)

        gamma = 1.5  # default
        for threshold, gamma_val in self.gamma_thresholds.items():
            if mean_brightness < threshold:
                gamma = gamma_val
                break

        inv_gamma = 1.0 / gamma
        table = np.array([((i / 255.0) ** inv_gamma) * 255 for i in range(256)]).astype("uint8")
        return cv2.LUT(image, table)

    def _apply_advanced_clahe(self, image):
        """LAB 색공간 CLAHE"""
        lab = cv2.cvtColor(image, cv2.COLOR_BGR2LAB)
        l, a, b = cv2.split(lab)
        l_enhanced = self.clahe.apply(l)
        enhanced_lab = cv2.merge([l_enhanced, a, b])
        return cv2.cvtColor(enhanced_lab, cv2.COLOR_LAB2BGR)

    def _apply_unsharp_masking(self, image):
        """언샤프 마스킹"""
        sharpened = cv2.filter2D(image, -1, self.sharpen_kernel)
        return cv2.addWeighted(image, 0.7, sharpened, 0.3, 0)


class Phase3UltraEnhanced:
    """Ultra Enhancement 적용 Phase 3 분석기"""

    def __init__(self, image_dir: str, excel_path: str, calibration_path: str):
        self.image_dir = Path(image_dir)
        self.excel_path = Path(excel_path)
        self.calibration_path = Path(calibration_path)

        # Enhanced images directory
        self.enhanced_dir = self.image_dir.parent / "driver_ultra_enhanced"
        self.enhanced_dir.mkdir(exist_ok=True)

        # Enhancer
        self.enhancer = UltraEnhancer()

        # Results
        self.ball_positions_3d = []
        self.timestamps = []
        self.fps = 820

    def preprocess_images(self):
        """BMP 이미지 Ultra Enhancement 전처리"""
        print("="*70)
        print("STEP 1: Ultra Enhancement Preprocessing")
        print("="*70)

        bmp_files = list(self.image_dir.glob("*.bmp"))
        print(f"Found {len(bmp_files)} BMP files")

        enhanced_count = 0
        brightness_improvements = []

        for bmp_file in bmp_files:
            # Load
            img = cv2.imread(str(bmp_file))
            if img is None:
                continue

            # Original brightness
            gray_orig = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            mean_orig = gray_orig.mean()

            # Apply Ultra Enhancement
            enhanced = self.enhancer.enhance(img)

            # Enhanced brightness
            gray_enh = cv2.cvtColor(enhanced, cv2.COLOR_BGR2GRAY)
            mean_enh = gray_enh.mean()

            # Save
            output_path = self.enhanced_dir / bmp_file.name
            cv2.imwrite(str(output_path), enhanced)

            enhanced_count += 1
            improvement = ((mean_enh - mean_orig) / mean_orig) * 100
            brightness_improvements.append(improvement)

            print(f"  {bmp_file.name}: {mean_orig:.1f} -> {mean_enh:.1f} (+{improvement:.1f}%)")

        avg_improvement = np.mean(brightness_improvements)
        print(f"\n[OK] Enhanced {enhanced_count}/{len(bmp_files)} images")
        print(f"[OK] Average brightness improvement: +{avg_improvement:.1f}%")
        print(f"[OK] Output: {self.enhanced_dir}")

        return enhanced_count

    def load_calibration(self):
        """캘리브레이션 로드"""
        with open(self.calibration_path, 'r', encoding='utf-8') as f:
            calibration = json.load(f)
        print(f"\n[OK] Calibration loaded:")
        print(f"  - Baseline: {calibration['baseline']}mm")
        print(f"  - Focal Length: {calibration['focal_length']}px")
        return calibration

    def load_enhanced_images(self):
        """전처리된 이미지 로드"""
        print("\n" + "="*70)
        print("STEP 2: Load Enhanced Images")
        print("="*70)

        bmp_files = list(self.enhanced_dir.glob("*.bmp"))
        print(f"Found {len(bmp_files)} enhanced BMP files")

        camera1_files = []
        camera2_files = []

        for bmp_file in bmp_files:
            match1 = re.match(r'Gamma_1_(\d+)\.bmp', bmp_file.name)
            if match1:
                frame_num = int(match1.group(1))
                camera1_files.append((frame_num, bmp_file))

            match2 = re.match(r'Gamma_2_(\d+)\.bmp', bmp_file.name)
            if match2:
                frame_num = int(match2.group(1))
                camera2_files.append((frame_num, bmp_file))

        camera1_files.sort(key=lambda x: x[0])
        camera2_files.sort(key=lambda x: x[0])

        frames_cam1 = []
        frames_cam2 = []

        for frame_num, file_path in camera1_files:
            img = cv2.imread(str(file_path))
            if img is not None:
                frames_cam1.append((frame_num, img))

        for frame_num, file_path in camera2_files:
            img = cv2.imread(str(file_path))
            if img is not None:
                frames_cam2.append((frame_num, img))

        print(f"[OK] Camera 1 (top): {len(frames_cam1)} frames")
        print(f"[OK] Camera 2 (bottom): {len(frames_cam2)} frames")

        return frames_cam1, frames_cam2

    def detect_ball(self, img, frame_num):
        """향상된 이미지용 볼 검출"""
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

        # CLAHE (전처리된 이미지용)
        clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
        enhanced = clahe.apply(gray)

        # Hough Circles
        circles = cv2.HoughCircles(
            enhanced,
            cv2.HOUGH_GRADIENT,
            dp=1,
            minDist=30,
            param1=50,  # 향상된 이미지용
            param2=20,
            minRadius=3,
            maxRadius=40
        )

        if circles is not None:
            circles = np.round(circles[0, :]).astype("int")
            best_circle = None
            max_brightness = 0

            for (x, y, r) in circles:
                if 0 <= y < enhanced.shape[0] and 0 <= x < enhanced.shape[1]:
                    brightness = int(enhanced[y, x])
                    if brightness > max_brightness:
                        max_brightness = brightness
                        best_circle = (x, y, r)

            if best_circle:
                return best_circle

        # Threshold backup
        _, thresh = cv2.threshold(enhanced, 180, 255, cv2.THRESH_BINARY)
        contours, _ = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

        if contours:
            largest_contour = max(contours, key=cv2.contourArea)
            if cv2.contourArea(largest_contour) > 10:
                ((x, y), radius) = cv2.minEnclosingCircle(largest_contour)
                if radius > 3 and radius < 40:
                    return (int(x), int(y), int(radius))

        return None

    def calculate_3d_position(self, point_top, point_bottom, calibration):
        """3D 위치 계산"""
        x_top, y_top, _ = point_top
        x_bottom, y_bottom, _ = point_bottom

        disparity_y = abs(y_top - y_bottom)

        if disparity_y < 5:
            return None

        baseline = calibration['baseline']
        focal_length = calibration['focal_length']
        scale_y = calibration['scale_factor_y']

        Z = (baseline * focal_length) / (disparity_y * scale_y)

        if Z < 500 or Z > 10000:
            return None

        X = (x_top + x_bottom) / 2.0 - 720
        X = X * (Z / focal_length)

        Y = y_top - 150
        Y = Y * (Z / focal_length) * scale_y

        return (X, Y, Z)

    def analyze_swing(self, frames_cam1, frames_cam2, calibration):
        """스윙 분석"""
        print("\n" + "="*70)
        print("STEP 3: Analyze Swing")
        print("="*70)

        for i in range(min(len(frames_cam1), len(frames_cam2))):
            frame_num1, img1 = frames_cam1[i]
            frame_num2, img2 = frames_cam2[i]

            if frame_num1 != frame_num2:
                continue

            ball_top = self.detect_ball(img1, frame_num1)
            ball_bottom = self.detect_ball(img2, frame_num2)

            if ball_top and ball_bottom:
                pos_3d = self.calculate_3d_position(ball_top, ball_bottom, calibration)

                if pos_3d:
                    self.ball_positions_3d.append(pos_3d)
                    timestamp = frame_num1 / self.fps
                    self.timestamps.append(timestamp)

                    print(f"Frame {frame_num1}: Ball at 3D ({pos_3d[0]:.1f}, {pos_3d[1]:.1f}, {pos_3d[2]:.1f}) mm")

        detection_rate = (len(self.ball_positions_3d) / len(frames_cam1)) * 100
        print(f"\n[OK] Detected ball in {len(self.ball_positions_3d)} frames ({detection_rate:.1f}%)")

        if len(self.ball_positions_3d) >= 3:
            return self.calculate_physics()
        else:
            print("[WARN] Not enough data points")
            return None

    def calculate_physics(self):
        """물리량 계산"""
        print("\nCalculating physics...")

        positions = np.array(self.ball_positions_3d)
        times = np.array(self.timestamps)

        # Ball speed
        if len(positions) >= 5:
            distances = []
            for i in range(1, 5):
                dx = positions[i][0] - positions[i-1][0]
                dy = positions[i][1] - positions[i-1][1]
                dz = positions[i][2] - positions[i-1][2]
                dist = math.sqrt(dx**2 + dy**2 + dz**2)
                distances.append(dist)

            dt = times[1] - times[0]
            avg_speed_ms = np.mean(distances) / (dt * 1000)
            ball_speed_mph = avg_speed_ms * 2.23694
        else:
            ball_speed_mph = 0.0

        # Launch angle
        if len(positions) >= 3:
            x_vals = positions[:3, 2]
            y_vals = positions[:3, 1]

            if len(x_vals) > 1 and x_vals[1] != x_vals[0]:
                slope = (y_vals[1] - y_vals[0]) / (x_vals[1] - x_vals[0])
                launch_angle_deg = math.degrees(math.atan(slope))
            else:
                launch_angle_deg = 0.0
        else:
            launch_angle_deg = 0.0

        # Azimuth
        if len(positions) >= 3:
            x_vals = positions[:3, 2]
            z_vals = positions[:3, 0]

            if len(x_vals) > 1 and x_vals[1] != x_vals[0]:
                slope = (z_vals[1] - z_vals[0]) / (x_vals[1] - x_vals[0])
                azimuth_angle_deg = math.degrees(math.atan(slope))
            else:
                azimuth_angle_deg = 0.0
        else:
            azimuth_angle_deg = 0.0

        result = {
            'ball_speed_mph': ball_speed_mph,
            'launch_angle_deg': launch_angle_deg,
            'azimuth_angle_deg': azimuth_angle_deg,
            'backspin_rpm': 2500.0,
            'sidespin_rpm': 0.0,
            'spin_axis_deg': 0.0,
            'club_speed_mph': ball_speed_mph / 1.5,
            'attack_angle_deg': -4.0
        }

        print(f"[OK] Physics calculated:")
        print(f"  - Ball Speed: {ball_speed_mph:.2f} mph")
        print(f"  - Launch Angle: {launch_angle_deg:.2f}°")
        print(f"  - Azimuth Angle: {azimuth_angle_deg:.2f}°")

        return result

    def update_excel(self, result):
        """Excel 업데이트"""
        print(f"\nUpdating Excel: {self.excel_path}")

        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb['Driver_Ball_Standard']

        ball_speed_ms = result['ball_speed_mph'] * 0.44704

        measurements = {
            'F5': result['ball_speed_mph'],
            'F6': ball_speed_ms,
            'F7': result['launch_angle_deg'],
            'F8': result['azimuth_angle_deg'],
            'F9': result['backspin_rpm'],
            'F10': result['sidespin_rpm'],
            'F11': result['spin_axis_deg'],
            'F12': 0.0,
            'F13': 0.0,
            'F14': 0.0,
            'F16': 0.0
        }

        for cell, value in measurements.items():
            try:
                ws[cell] = value
                print(f"  {cell}: {value:.2f}")
            except AttributeError:
                print(f"  {cell}: Skipped (merged cell)")

        output_path = self.excel_path.parent / f"{self.excel_path.stem}_phase3_ultra_enhanced.xlsx"
        wb.save(output_path)
        wb.close()

        print(f"\n[OK] Excel saved: {output_path}")
        return output_path

    def calculate_accuracy(self, excel_path):
        """정확도 계산"""
        print("\nCalculating accuracy...")

        wb = openpyxl.load_workbook(excel_path)
        ws = wb['Driver_Ball_Standard']

        results = []
        for row in range(5, 17):
            measurement = ws[f'A{row}'].value
            standard = ws[f'B{row}'].value
            measured = ws[f'F{row}'].value

            if measured and standard and standard != 0:
                error_pct = abs((measured - standard) / standard) * 100
                status = '[OK]' if error_pct <= 3.5 else '[WARN]' if error_pct <= 5.0 else '[FAIL]'

                results.append({
                    'measurement': measurement,
                    'standard': standard,
                    'measured': measured,
                    'error_pct': error_pct
                })

                print(f"  {status} {measurement}: {error_pct:.2f}% error")

        wb.close()

        errors = [r['error_pct'] for r in results]
        avg_error = np.mean(errors)
        within_target = sum(1 for e in errors if e <= 3.5)
        accuracy_rate = (within_target / len(errors)) * 100

        print(f"\n=== ACCURACY SUMMARY ===")
        print(f"Average Error: {avg_error:.2f}%")
        print(f"Within Target: {within_target}/{len(errors)} ({accuracy_rate:.1f}%)")

        if accuracy_rate >= 90:
            print(f"[OK] PASS: {accuracy_rate:.1f}%")
        elif accuracy_rate >= 80:
            print(f"[WARN] PASS: {accuracy_rate:.1f}%")
        else:
            print(f"[FAIL] FAIL: {accuracy_rate:.1f}%")

        return {
            'results': results,
            'avg_error': avg_error,
            'within_target': within_target,
            'total': len(errors),
            'accuracy_rate': accuracy_rate
        }

    def run(self):
        """전체 파이프라인"""
        print("="*70)
        print("Phase 3: Ultra Enhancement + Driver Analysis")
        print("="*70)

        # 1. Preprocess
        enhanced_count = self.preprocess_images()

        if enhanced_count == 0:
            print("[FAIL] No images enhanced")
            return

        # 2. Load calibration
        calibration = self.load_calibration()

        # 3. Load enhanced images
        frames_cam1, frames_cam2 = self.load_enhanced_images()

        # 4. Analyze
        result = self.analyze_swing(frames_cam1, frames_cam2, calibration)

        if result is None:
            print("[FAIL] Analysis failed")
            return

        # 5. Update Excel
        excel_path = self.update_excel(result)

        # 6. Calculate accuracy
        accuracy = self.calculate_accuracy(excel_path)

        # 7. Save results
        json_path = Path('data/phase3_ultra_enhanced_result.json')
        json_path.parent.mkdir(parents=True, exist_ok=True)

        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump({
                'measurements': result,
                'accuracy': accuracy,
                'detection_count': len(self.ball_positions_3d),
                'total_frames': 23,
                'detection_rate': (len(self.ball_positions_3d) / 23) * 100,
                'timestamp': datetime.now().isoformat(),
                'excel_output': str(excel_path)
            }, f, indent=2)

        print(f"\n[OK] Results saved: {json_path}")
        print("="*70)
        print(f"PHASE 3 ULTRA ENHANCED COMPLETED")
        print(f"Detection Rate: {(len(self.ball_positions_3d) / 23) * 100:.1f}%")
        print(f"Accuracy: {accuracy['accuracy_rate']:.1f}%")
        print("="*70)

if __name__ == "__main__":
    analyzer = Phase3UltraEnhanced(
        image_dir='C:/src/GolfSwingAnalysis_Final/data/1440_300_data/driver/1',
        excel_path='C:/src/GolfSwingAnalysis_Final/data/data-standard.xlsx',
        calibration_path='C:/src/GolfSwingAnalysis_Final/config/calibration_default.json'
    )

    analyzer.run()

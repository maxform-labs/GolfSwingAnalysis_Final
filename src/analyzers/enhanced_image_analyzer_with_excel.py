"""
Enhanced Image Analyzer with Excel Report Generation
Analyzes golf ball images frame by frame and generates detailed Excel reports with embedded images
"""

import cv2
import numpy as np
import pandas as pd
from pathlib import Path
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from PIL import Image, ImageEnhance
import io
import matplotlib.pyplot as plt
from typing import Dict, List, Tuple, Optional
import warnings
warnings.filterwarnings('ignore')

class EnhancedImageAnalyzer:
    """Enhanced image analyzer with multiple detection methods and Excel reporting"""
    
    def __init__(self, base_path: str = "C:/src/GolfSwingAnalysis_Final_ver8"):
        self.base_path = Path(base_path)
        self.results_dir = self.base_path / "analysis_results"
        self.results_dir.mkdir(exist_ok=True)
        
        # Detection parameters optimized for golf balls
        self.params = {
            'motion_threshold': 15,
            'min_area': 50,
            'max_area': 5000,
            'circularity_threshold': 0.5,
            'dp': 1.2,
            'minDist': 30,
            'param1': 50,
            'param2': 30,
            'minRadius': 5,
            'maxRadius': 50
        }
        
    def enhance_image(self, img: np.ndarray) -> Dict[str, np.ndarray]:
        """Apply multiple enhancement techniques to improve ball detection"""
        enhanced = {}
        
        # Convert to PIL for enhancement
        pil_img = Image.fromarray(cv2.cvtColor(img, cv2.COLOR_BGR2RGB))
        
        # 1. Brightness enhancement
        brightness_enhancer = ImageEnhance.Brightness(pil_img)
        bright_img = brightness_enhancer.enhance(2.0)
        enhanced['brightness'] = cv2.cvtColor(np.array(bright_img), cv2.COLOR_RGB2BGR)
        
        # 2. Contrast enhancement
        contrast_enhancer = ImageEnhance.Contrast(pil_img)
        contrast_img = contrast_enhancer.enhance(2.5)
        enhanced['contrast'] = cv2.cvtColor(np.array(contrast_img), cv2.COLOR_RGB2BGR)
        
        # 3. Gamma correction
        gamma = 2.0
        inv_gamma = 1.0 / gamma
        table = np.array([((i / 255.0) ** inv_gamma) * 255 for i in np.arange(0, 256)]).astype("uint8")
        enhanced['gamma'] = cv2.LUT(img, table)
        
        # 4. CLAHE (Contrast Limited Adaptive Histogram Equalization)
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        clahe = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8,8))
        clahe_img = clahe.apply(gray)
        enhanced['clahe'] = cv2.cvtColor(clahe_img, cv2.COLOR_GRAY2BGR)
        
        # 5. Adaptive threshold
        adaptive = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                       cv2.THRESH_BINARY, 11, 2)
        enhanced['adaptive'] = cv2.cvtColor(adaptive, cv2.COLOR_GRAY2BGR)
        
        return enhanced
    
    def detect_ball_motion(self, img: np.ndarray, prev_img: Optional[np.ndarray] = None) -> List[Dict]:
        """Detect ball using motion detection"""
        if prev_img is None:
            return []
        
        # Frame differencing
        diff = cv2.absdiff(img, prev_img)
        gray_diff = cv2.cvtColor(diff, cv2.COLOR_BGR2GRAY)
        
        # Threshold
        _, thresh = cv2.threshold(gray_diff, self.params['motion_threshold'], 255, cv2.THRESH_BINARY)
        
        # Morphological operations
        kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (5, 5))
        thresh = cv2.morphologyEx(thresh, cv2.MORPH_CLOSE, kernel)
        thresh = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, kernel)
        
        # Find contours
        contours, _ = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        
        detections = []
        for contour in contours:
            area = cv2.contourArea(contour)
            if self.params['min_area'] < area < self.params['max_area']:
                # Check circularity
                perimeter = cv2.arcLength(contour, True)
                if perimeter > 0:
                    circularity = 4 * np.pi * area / (perimeter * perimeter)
                    if circularity > self.params['circularity_threshold']:
                        (x, y), radius = cv2.minEnclosingCircle(contour)
                        detections.append({
                            'method': 'motion',
                            'x': int(x),
                            'y': int(y),
                            'radius': int(radius),
                            'confidence': circularity,
                            'area': area
                        })
        
        return detections
    
    def detect_ball_hough(self, img: np.ndarray, enhanced_type: str = 'original') -> List[Dict]:
        """Detect ball using Hough Circle Transform"""
        if len(img.shape) == 3:
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        else:
            gray = img
        
        # Apply Gaussian blur
        blurred = cv2.GaussianBlur(gray, (9, 9), 2)
        
        # Hough Circle Transform
        circles = cv2.HoughCircles(
            blurred,
            cv2.HOUGH_GRADIENT,
            dp=self.params['dp'],
            minDist=self.params['minDist'],
            param1=self.params['param1'],
            param2=self.params['param2'],
            minRadius=self.params['minRadius'],
            maxRadius=self.params['maxRadius']
        )
        
        detections = []
        if circles is not None:
            circles = np.uint16(np.around(circles))
            for circle in circles[0, :]:
                detections.append({
                    'method': f'hough_{enhanced_type}',
                    'x': int(circle[0]),
                    'y': int(circle[1]),
                    'radius': int(circle[2]),
                    'confidence': 0.8,
                    'area': np.pi * circle[2] * circle[2]
                })
        
        return detections
    
    def detect_ball_contours(self, img: np.ndarray) -> List[Dict]:
        """Detect ball using contour detection"""
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY) if len(img.shape) == 3 else img
        
        # Apply threshold
        _, binary = cv2.threshold(gray, 127, 255, cv2.THRESH_BINARY)
        
        # Find contours
        contours, _ = cv2.findContours(binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        
        detections = []
        for contour in contours:
            area = cv2.contourArea(contour)
            if self.params['min_area'] < area < self.params['max_area']:
                # Fit circle
                (x, y), radius = cv2.minEnclosingCircle(contour)
                
                # Check if contour is circular
                perimeter = cv2.arcLength(contour, True)
                if perimeter > 0:
                    circularity = 4 * np.pi * area / (perimeter * perimeter)
                    if circularity > self.params['circularity_threshold']:
                        detections.append({
                            'method': 'contour',
                            'x': int(x),
                            'y': int(y),
                            'radius': int(radius),
                            'confidence': circularity,
                            'area': area
                        })
        
        return detections
    
    def calculate_ball_properties(self, detection: Dict, frame_num: int, fps: int = 820) -> Dict:
        """Calculate ball properties from detection"""
        properties = {
            'frame_number': frame_num,
            'x_position': detection['x'],
            'y_position': detection['y'],
            'radius': detection['radius'],
            'diameter_pixels': detection['radius'] * 2,
            'area_pixels': detection['area'],
            'detection_method': detection['method'],
            'confidence': detection['confidence'],
            'timestamp_ms': (frame_num / fps) * 1000
        }
        
        # Estimate physical properties (assuming known ball diameter of 42.67mm)
        ball_diameter_mm = 42.67
        if detection['radius'] > 0:
            pixels_per_mm = (detection['radius'] * 2) / ball_diameter_mm
            properties['pixels_per_mm'] = pixels_per_mm
            
            # Calculate physical area
            properties['physical_area_mm2'] = np.pi * (ball_diameter_mm / 2) ** 2
        
        return properties
    
    def draw_detection(self, img: np.ndarray, detections: List[Dict], frame_num: int) -> np.ndarray:
        """Draw detection results on image"""
        result_img = img.copy()
        
        for det in detections:
            # Draw circle
            cv2.circle(result_img, (det['x'], det['y']), det['radius'], (0, 255, 0), 2)
            cv2.circle(result_img, (det['x'], det['y']), 2, (0, 0, 255), -1)
            
            # Add text
            text = f"{det['method']} ({det['confidence']:.2f})"
            cv2.putText(result_img, text, (det['x'] - 30, det['y'] - det['radius'] - 10),
                       cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 0), 1)
        
        # Add frame number
        cv2.putText(result_img, f"Frame: {frame_num}", (10, 30),
                   cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)
        
        return result_img
    
    def analyze_folder(self, folder_name: str) -> pd.DataFrame:
        """Analyze all images in a folder"""
        folder_path = self.base_path / "shot-image" / folder_name
        
        if not folder_path.exists():
            # Try alternative path
            folder_path = self.base_path / "shot-image" / "7iron" / folder_name
            if not folder_path.exists():
                print(f"Folder not found: {folder_name}")
                return pd.DataFrame()
        
        # Get all camera 1 images
        image_files = sorted(folder_path.glob("1_*.bmp"))
        if not image_files:
            image_files = sorted(folder_path.glob("1_*.jpg"))
        
        print(f"Found {len(image_files)} images in {folder_name}")
        
        results = []
        processed_images = {}
        prev_img = None
        
        for idx, img_path in enumerate(image_files):
            # Extract frame number from filename
            frame_num = int(img_path.stem.split('_')[1])
            
            # Read image
            img = cv2.imread(str(img_path))
            if img is None:
                continue
            
            # Get enhanced versions
            enhanced_versions = self.enhance_image(img)
            
            # Detect ball using multiple methods
            all_detections = []
            
            # 1. Motion detection
            if prev_img is not None:
                motion_detections = self.detect_ball_motion(img, prev_img)
                all_detections.extend(motion_detections)
            
            # 2. Hough transform on original
            hough_detections = self.detect_ball_hough(img, 'original')
            all_detections.extend(hough_detections)
            
            # 3. Hough transform on enhanced versions
            for enhance_type, enhanced_img in enhanced_versions.items():
                detections = self.detect_ball_hough(enhanced_img, enhance_type)
                all_detections.extend(detections)
            
            # 4. Contour detection
            contour_detections = self.detect_ball_contours(enhanced_versions['adaptive'])
            all_detections.extend(contour_detections)
            
            # Select best detection
            if all_detections:
                best_detection = max(all_detections, key=lambda x: x['confidence'])
                
                # Calculate properties
                properties = self.calculate_ball_properties(best_detection, frame_num)
                
                # Draw detection
                result_img = self.draw_detection(img, [best_detection], frame_num)
                
                # Store processed image
                processed_images[frame_num] = result_img
                
                results.append(properties)
            else:
                # No detection
                results.append({
                    'frame_number': frame_num,
                    'x_position': None,
                    'y_position': None,
                    'radius': None,
                    'detection_method': 'none',
                    'confidence': 0
                })
                processed_images[frame_num] = img
            
            prev_img = img
        
        # Create DataFrame
        df = pd.DataFrame(results)
        
        # Save to Excel with images
        self.save_to_excel(df, processed_images, folder_name)
        
        return df
    
    def save_to_excel(self, df: pd.DataFrame, images: Dict[int, np.ndarray], folder_name: str):
        """Save results to Excel with embedded images"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_path = self.results_dir / f"{folder_name}_analysis_{timestamp}.xlsx"
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Analysis Results"
        
        # Set column widths
        ws.column_dimensions['A'].width = 15  # Frame Number
        ws.column_dimensions['B'].width = 12  # X Position
        ws.column_dimensions['C'].width = 12  # Y Position
        ws.column_dimensions['D'].width = 10  # Radius
        ws.column_dimensions['E'].width = 15  # Diameter
        ws.column_dimensions['F'].width = 15  # Area
        ws.column_dimensions['G'].width = 20  # Method
        ws.column_dimensions['H'].width = 12  # Confidence
        ws.column_dimensions['I'].width = 15  # Timestamp
        ws.column_dimensions['J'].width = 40  # Processed Image
        
        # Add headers with formatting
        headers = ['Frame Number', 'X Position', 'Y Position', 'Radius (px)', 
                  'Diameter (px)', 'Area (px²)', 'Detection Method', 
                  'Confidence', 'Timestamp (ms)', 'Processed Image']
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add data rows
        for idx, row in df.iterrows():
            row_num = idx + 2
            
            # Add data
            ws.cell(row=row_num, column=1, value=row.get('frame_number', ''))
            ws.cell(row=row_num, column=2, value=row.get('x_position', ''))
            ws.cell(row=row_num, column=3, value=row.get('y_position', ''))
            ws.cell(row=row_num, column=4, value=row.get('radius', ''))
            ws.cell(row=row_num, column=5, value=row.get('diameter_pixels', ''))
            ws.cell(row=row_num, column=6, value=row.get('area_pixels', ''))
            ws.cell(row=row_num, column=7, value=row.get('detection_method', ''))
            ws.cell(row=row_num, column=8, value=row.get('confidence', ''))
            ws.cell(row=row_num, column=9, value=row.get('timestamp_ms', ''))
            
            # Add processed image
            frame_num = row.get('frame_number')
            if frame_num in images:
                # Convert image to bytes
                img = images[frame_num]
                img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                pil_img = Image.fromarray(img_rgb)
                
                # Resize image to fit in cell
                max_width = 300
                max_height = 200
                pil_img.thumbnail((max_width, max_height), Image.Resampling.LANCZOS)
                
                # Save to bytes
                img_bytes = io.BytesIO()
                pil_img.save(img_bytes, format='PNG')
                img_bytes.seek(0)
                
                # Add to Excel
                xl_img = XLImage(img_bytes)
                xl_img.anchor = f'J{row_num}'
                ws.add_image(xl_img)
                
                # Set row height to accommodate image
                ws.row_dimensions[row_num].height = max_height * 0.75
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=len(df)+1, min_col=1, max_col=10):
            for cell in row:
                cell.border = thin_border
                if cell.row > 1:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add summary sheet
        ws2 = wb.create_sheet("Summary")
        
        # Summary statistics
        summary_data = [
            ["Metric", "Value"],
            ["Total Frames", len(df)],
            ["Detected Frames", len(df[df['confidence'] > 0])],
            ["Detection Rate", f"{len(df[df['confidence'] > 0]) / len(df) * 100:.1f}%"],
            ["Average Confidence", f"{df['confidence'].mean():.3f}"],
            ["Most Common Method", df['detection_method'].mode()[0] if len(df['detection_method'].mode()) > 0 else 'N/A']
        ]
        
        for row_idx, row_data in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        ws2.column_dimensions['A'].width = 20
        ws2.column_dimensions['B'].width = 20
        
        # Save workbook
        wb.save(excel_path)
        print(f"Results saved to: {excel_path}")
        
        return excel_path

def main():
    """Main function to analyze 7iron_logo_ball-1"""
    analyzer = EnhancedImageAnalyzer()
    
    # Analyze the specific folder
    folder_name = "logo_ball-1"  # Using English path naming
    print(f"Analyzing folder: {folder_name}")
    
    df = analyzer.analyze_folder(folder_name)
    
    if not df.empty:
        print("\n=== Analysis Summary ===")
        print(f"Total frames analyzed: {len(df)}")
        print(f"Frames with detection: {len(df[df['confidence'] > 0])}")
        print(f"Detection rate: {len(df[df['confidence'] > 0]) / len(df) * 100:.1f}%")
        print(f"Average confidence: {df['confidence'].mean():.3f}")
        
        # Show detection methods used
        method_counts = df['detection_method'].value_counts()
        print("\nDetection methods:")
        for method, count in method_counts.items():
            print(f"  {method}: {count} frames")
        
        # Calculate ball speed if positions are available
        valid_positions = df[df['x_position'].notna()].copy()
        if len(valid_positions) > 1:
            # Calculate speed between consecutive frames
            valid_positions['dx'] = valid_positions['x_position'].diff()
            valid_positions['dy'] = valid_positions['y_position'].diff()
            valid_positions['dt'] = valid_positions['timestamp_ms'].diff() / 1000  # Convert to seconds
            
            valid_positions['speed_pixels_per_sec'] = np.sqrt(
                valid_positions['dx']**2 + valid_positions['dy']**2
            ) / valid_positions['dt']
            
            # Convert to approximate mph (assuming pixel to meter conversion)
            # This is a rough estimate - actual conversion requires calibration
            pixels_per_meter = 1000  # Approximate value
            valid_positions['speed_mps'] = valid_positions['speed_pixels_per_sec'] / pixels_per_meter
            valid_positions['speed_mph'] = valid_positions['speed_mps'] * 2.237
            
            avg_speed = valid_positions['speed_mph'].mean()
            max_speed = valid_positions['speed_mph'].max()
            
            print(f"\nEstimated ball speed:")
            print(f"  Average: {avg_speed:.1f} mph")
            print(f"  Maximum: {max_speed:.1f} mph")
    else:
        print("No data analyzed")

if __name__ == "__main__":
    main()